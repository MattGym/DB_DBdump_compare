from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


def get_col_no(sheet, desc, max_col):
    """
    According to given active_sheet name and searched column description (str), function
    returns column number where that string is.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    desc : str
        searched string
    max_col : int
        number of not empty columns
    """
    for s in range(max_col):
        if str(desc) == str(sheet.cell(row=1, column=s+1).value):
            return s+1
    return 0


def set_cell_color(sheet, row, col, color_rgn='n'):
    """
    Change cell background color on 'r'-red, 'g'-green, 'n'-none.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    color_rgn : str, optional
        (OPTIONAL) color_rgn = 'r'-red; 'g'-green; 'n' or empty - NONE
    """
    if color_rgn == 'r':
        fg_color = 'E74C3C'
        pattern = PatternFill(patternType='solid', fgColor=fg_color)
        sheet.cell(row=row, column=col).fill = pattern
    if color_rgn == 'g':
        fg_color = '2ECC71'
        pattern = PatternFill(patternType='solid', fgColor=fg_color)
        sheet.cell(row=row, column=col).fill = pattern
    if color_rgn == 'n':
        pattern = PatternFill(patternType=None)
        sheet.cell(row=row, column=col).fill = pattern


def get_cell_value(sheet, row, col):
    """
    Returns value from cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    """
    return sheet.cell(row=row, column=col).value


def set_cell_value(sheet, row, col, val, typ=0):
    """
    Set value at cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    val : any
        value
    typ : int, optional
        (OPTIONAL) type of data (0 or empty - float/int; 1 - force string; 2 - None)
    """
    if typ == 0:
        sheet.cell(row=row, column=col).value = float(val)
    elif typ == 1:
        sheet.cell(row=row, column=col).value = str(val)
    elif typ == 2:
        sheet.cell(row=row, column=col).value = None
    elif typ == 3:
        sheet.cell(row=row, column=col).value = int(val)


def set_cell_comment(sheet, row, col, commentary, add=False, delete=False):
    """
    Function add 'commentary' to specified cell given as (active_sheet, row & column position)
    If optional parameter add=True then function add commentary to existing one.
    In another way removes old commentary and add a new one.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    row : int
        cell row position
    col : int
        cell column position
    commentary : str
        commentary that will be added to specified cell
    add : bool, optional
        (OPTIONAL) add=False or NONE - function swap commentary into new on , add=True - add second commentary
    delete : bool, optional
        (OPTIONAL) delete=False or NONE - function do nothing with existing commentary, delete=True - remove commentary
    """
    if str(sheet.cell(row=row, column=col).comment) == 'None' or add is False:
        comment = Comment(commentary, 'CdA analyzer')
        comment.width = 400
        comment.height = 150
        sheet.cell(row=row, column=col).comment = comment
    else:
        tmp_txt1 = str(str(sheet.cell(row=row, column=col).comment).replace('Comment: ', '')).\
            replace('by CdA analyzer', '')
        comment = Comment(tmp_txt1 + ' ::\n' + commentary, 'CdA analyzer')
        comment.width = 400
        comment.height = 150
        sheet.cell(row=row, column=col).comment = comment
    if delete is True:
        sheet.cell(row=row, column=col).comment = None


def remove_sign(sheet, col, max_row, sign=' '):
    """
    Returns sign form cell in active_sheet at specific row & column position.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    col : list
        column no to be changed
    max_row : int
        number of rows in active_sheet
    sign : str, optional
        (OPTIONAL) sign to be removed (space is default)
    """
    for i in col:
        for j in range(2, max_row+1):
            if get_cell_value(sheet, j, i) is not None:
                set_cell_value(sheet, j, i, str(sheet.cell(j, i).value).replace(sign, ''), 1)
                if get_cell_value(sheet, j, i) == ' ' or get_cell_value(sheet, j, i) == '  ' or \
                        get_cell_value(sheet, j, i) == '   ' or get_cell_value(sheet, j, i) == '   ':
                    set_cell_value(sheet, j, i, '')


def variable_type_update(sheet, col_v, col_s, col_txt, max_row):
    """
    Function change  data types in cells for specific columns.
    Parameters
    ----------
    sheet : workbook.active
        active sheet name
    col_v : list
        list of columns to be updated as value/float
    col_s : list
        list of number columns to be updated as string
    col_txt : list
        list of columns as string
    max_row : int
        number of rows in active_sheet
    """
    for i in col_v:
        mark_col = False
        for j in range(2, max_row+1):
            if get_cell_value(sheet, j, i) == '':
                set_cell_value(sheet, j, i, 0, 2)
            if get_cell_value(sheet, j, i) is not None:
                try:
                    temp_val = float(get_cell_value(sheet, j, i))
                    set_cell_value(sheet, j, i, temp_val)
                except:
                    if get_cell_value(sheet, j, i) is not None and get_cell_value(sheet, j, i) != '':
                        set_cell_color(sheet, j, i, 'r')
                        set_cell_comment(sheet, j, i, 'Wrong data type it has to be a number or nothing')
                        mark_col = True
        if mark_col:
            set_cell_color(sheet, 1, i, 'r')
    for i in col_s:
        for j in range(2, max_row+1):
            if get_cell_value(sheet, j, i) is not None:
                if len(str(get_cell_value(sheet, j, i))) == 1:
                    set_cell_value(sheet, j, i, '0' + str(get_cell_value(sheet, j, i)), 1)
                    set_cell_color(sheet, j, i, 'g')
                    set_cell_comment(sheet, j, i, 'Added 0 as prefix')
                else:
                    set_cell_value(sheet, j, i, str(get_cell_value(sheet, j, i)), 1)
    for i in col_txt:
        for j in range(2, max_row+1):
            if get_cell_value(sheet, j, i) is not None:
                if get_cell_value(sheet, j, i) == '':
                    set_cell_value(sheet, j, i, 0, 2)


def search_row_in_dump(sheet1, sheet2, no, file1_col_tag, file2_col_tag, file2_col_seq, max_rows1):
    """
    Function 'search same row' returns number of other row with same cells value in given columns.
    If function returns 0 that is means there is no duplicate
    Parameters
    ----------
    sheet1 : workbook.active
        active sheet name DUMP from DNA
    sheet2 : workbook.active
        active sheet name CdA DB
    no : int
        row number of searched row from CdA DB
    file1_col_tag : int
        column number with TAG name in DUMP from DNA
    file2_col_tag : int
        column number with TAG name in DB from CdA
    file2_col_seq : int
        column number with SEQ no in DB from CdA
    max_rows1
        numbers of rows in DUMP from CdA
    """
    row_found = 0
    loop_name = get_cell_value(sheet2, no, file2_col_tag) + '.' + str(int(get_cell_value(sheet2, no, file2_col_seq)))
    for i in range(2, max_rows1+1):
        if loop_name == str(get_cell_value(sheet1, i, file1_col_tag)):
            row_found = i
            return row_found
    return row_found


def search_row_in_db(sheet1, sheet2, no, file1_col_loop, file2_col_tag, max_rows2):
    print('dalej jechac z funkcja')


def compare_hw_address(sheet1, sheet2, file1_col_package, file2_col_package, file1_col_fbc, file2_col_fbc, file1_col_ibc
                       , file2_col_ibc, file1_col_card, file2_col_card, file1_col_channel, file2_col_channel
                       , file1_col_max, file2_col_max, file1_col_min, file2_col_min, file1_col_tag, file2_col_tag
                       , file2_col_io_type, file2_col_seq, max_rows1, max_rows2):
    """
    Function 'compare hw address' check if CdA db HW signals exist and are the same in DNA db dump.
    Function doesn't check TREAR signals in search of Alarm levels.
    Parameters
    ----------
    sheet1 : workbook
        active sheet name DUMP from DNA
    sheet2 : workbook
        active sheet name CdA DB
    file1_col_package : int
        column number for package in DNA DB dump
    file2_col_package : int
        column number for package in CdA DB
    file1_col_fbc : int
        column number for fbc in DNA DB dump
    file2_col_fbc : int
        column number for fbc in CdA DB
    file1_col_ibc : int
        column number for ibc in DNA DB dump
    file2_col_ibc : int
        column number for ibc in CdA DB
    file1_col_card : int
        column number for card in DNA DB dump
    file2_col_card : int
        column number for card in CdA DB
    file1_col_channel : int
        column number for channel in DNA DB dump
    file2_col_channel : int
        column number for channel in CdA DB
    file1_col_max : int
        column number for max in DNA DB dump
    file2_col_max : int
        column number for max in CdA DB
    file1_col_min : int
        column number for min in DNA DB dump
    file2_col_min : int
        column number for min in CdA DB
    file1_col_tag : int
        column number for tag in DNA DB dump
    file2_col_tag : int
        column number for tag in CdA DB
    file2_col_io_type : int
        column number for signal type in CdA DB
    file2_col_seq : int
        column number for sequence number in CdA DB
    max_min : int
        checkbox for checking max and minimum value for IO analog signals
    max_rows1 : int
        numbers of rows in DUMP from DNA DB dump
    max_rows2 : int
        numbers of rows in DUMP from CdA
    """
    mark_col_package = False
    mark_col_fbc = False
    mark_col_ibc = False
    mark_col_card = False
    mark_col_channel = False
    mark_col_max = False
    mark_col_min = False
    for i in range(2, max_rows2+1):
        no1 = 0
        if 'SL' not in str(get_cell_value(sheet2, i, file2_col_io_type)) \
                and 'TREAT' not in str(get_cell_value(sheet2, i, file2_col_io_type)):
            no1 = search_row_in_dump(sheet1, sheet2, i, file1_col_tag, file2_col_tag, file2_col_seq, max_rows1)
            if no1 != 0:
                if 'SL' not in str(get_cell_value(sheet2, i, file2_col_io_type)) \
                        and 'TREAT' not in str(get_cell_value(sheet2, i, file2_col_io_type)):
                    # package check
                    if str(get_cell_value(sheet1, no1, file1_col_package)) \
                            != str(get_cell_value(sheet2, i, file2_col_package)):
                        set_cell_color(sheet1, no1, file1_col_package, 'r')
                        set_cell_comment(sheet1, no1, file1_col_package, 'Other package than in CdA DB - ('
                                         + str(get_cell_value(sheet2, i, file2_col_package)), ')-ROW: ' + str(i)
                                         + '\n', True)
                        mark_col_package = True
                    # fbc check
                    if str(get_cell_value(sheet1, no1, file1_col_fbc)) \
                            != str(get_cell_value(sheet2, i, file2_col_fbc)):
                        set_cell_color(sheet1, no1, file1_col_fbc, 'r')
                        set_cell_comment(sheet1, no1, file1_col_fbc, 'Other FBC than in CdA DB - ('
                                         + str(get_cell_value(sheet2, i, file2_col_fbc)), ')-ROW: ' + str(i)
                                         + '\n', True)
                        mark_col_fbc = True
                    # ibc check
                    if str(get_cell_value(sheet1, no1, file1_col_ibc)) \
                            != str(get_cell_value(sheet2, i, file2_col_ibc)):
                        set_cell_color(sheet1, no1, file1_col_ibc, 'r')
                        set_cell_comment(sheet1, no1, file1_col_ibc, 'Other IBC than in CdA DB - ('
                                         + str(get_cell_value(sheet2, i, file2_col_ibc)), ')-ROW: ' + str(i)
                                         + '\n', True)
                        mark_col_ibc = True
                    # card check
                    if str(get_cell_value(sheet1, no1, file1_col_card)) \
                            != str(get_cell_value(sheet2, i, file2_col_card)):
                        set_cell_color(sheet1, no1, file1_col_card, 'r')
                        set_cell_comment(sheet1, no1, file1_col_card, 'Other CARD than in CdA DB - ('
                                         + str(get_cell_value(sheet2, i, file2_col_card)), ')-ROW: ' + str(i)
                                         + '\n', True)
                        mark_col_card = True
                    # channel check
                    if str(get_cell_value(sheet1, no1, file1_col_channel)) \
                            != str(get_cell_value(sheet2, i, file2_col_channel)):
                        set_cell_color(sheet1, no1, file1_col_channel, 'r')
                        set_cell_comment(sheet1, no1, file1_col_channel, 'Other CHANNEL than in CdA DB - ('
                                         + str(get_cell_value(sheet2, i, file2_col_channel)), ')-ROW: ' + str(i)
                                         + '\n', True)
                        mark_col_channel = True
                    if 'A' in str(get_cell_value(sheet2, i, file2_col_io_type)):
                        # max check (only for AI / AO)
                        if str(get_cell_value(sheet1, no1, file1_col_max)) \
                                != str(get_cell_value(sheet2, i, file2_col_max)):
                            set_cell_color(sheet1, no1, file1_col_max, 'r')
                            set_cell_comment(sheet1, no1, file1_col_max, 'Other MAX VALUE than in CdA DB - ('
                                             + str(get_cell_value(sheet2, i, file2_col_max)), ')-ROW: ' + str(i)
                                             + '\n', True)
                            mark_col_max = True
                        if str(get_cell_value(sheet1, no1, file1_col_min)) \
                                != str(get_cell_value(sheet2, i, file2_col_min)):
                            set_cell_color(sheet1, no1, file1_col_min, 'r')
                            set_cell_comment(sheet1, no1, file1_col_min, 'Other MIN VALUE than in CdA DB - ('
                                             + str(get_cell_value(sheet2, i, file2_col_min)), ')-ROW: ' + str(i)
                                             + '\n', True)
                            mark_col_min = True
            if no1 == 0:
                if 'SL' not in str(get_cell_value(sheet2, i, file2_col_io_type)) \
                        and 'TREAT' not in str(get_cell_value(sheet2, i, file2_col_io_type)):
                    set_cell_color(sheet2, i, 1, 'r')
                    set_cell_comment(sheet2, i, 1, 'Missing this HW signal as a "IO:child" in DNA db dump', True)
    if mark_col_package is True:
        set_cell_color(sheet1, 1, file1_col_package, 'r')
    if mark_col_fbc is True:
        set_cell_color(sheet1, 1, file1_col_fbc, 'r')
    if mark_col_ibc is True:
        set_cell_color(sheet1, 1, file1_col_ibc, 'r')
    if mark_col_card is True:
        set_cell_color(sheet1, 1, file1_col_card, 'r')
    if mark_col_channel is True:
        set_cell_color(sheet1, 1, file1_col_channel, 'r')
    if mark_col_max is True:
        set_cell_color(sheet1, 1, file1_col_max, 'r')
    if mark_col_min is True:
        set_cell_color(sheet1, 1, file1_col_min, 'r')


def compare_address(sheet1, sheet2, no1, no2, file1_col_package, file2_col_package, file1_col_fbc, file2_col_fbc,
                    file1_col_ibc, file2_col_ibc, file1_col_card, file2_col_card, file1_col_channel, file2_col_channel,
                    file2_col_iotype, file1_col_modbus_address, file2_col_modbus_address, file1_col_bit,
                    file2_col_bit, file1_col_gain, file2_col_gain, file1_col_slave, file2_col_slave):
    mark_row = False
    txt = 'Other than in CdA DB in row' + str(no2)
    if 'TREAT' not in get_cell_value(sheet2, no2, file2_col_iotype):
        if str(get_cell_value(sheet1, no1, file1_col_package)) != str(get_cell_value(sheet2, no2, file2_col_package)):
            set_cell_color(sheet1, no1, file1_col_package, 'r')
            set_cell_comment(sheet1, no1, file1_col_package, txt + ' in DB is ' +
                             str(get_cell_value(sheet2, no2, file2_col_package)), add=False)
            mark_row = True
        if 'SL' not in get_cell_value(sheet2, no2, file2_col_iotype):
            if int(get_cell_value(sheet1, no1, file1_col_fbc)) != int(get_cell_value(sheet2, no2, file2_col_fbc)):
                set_cell_color(sheet1, no1, file1_col_fbc, 'r')
                set_cell_comment(sheet1, no1, file1_col_fbc, txt + ' in DB is ' +
                                 str(int(get_cell_value(sheet2, no2, file2_col_fbc))), add=False)
                mark_row = True
            if int(get_cell_value(sheet1, no1, file1_col_ibc)) != int(get_cell_value(sheet2, no2, file2_col_ibc)):
                set_cell_color(sheet1, no1, file1_col_ibc, 'r')
                set_cell_comment(sheet1, no1, file1_col_ibc, txt + ' in DB is ' +
                                 str(int(get_cell_value(sheet2, no2, file2_col_ibc))), add=False)
                mark_row = True
            if int(get_cell_value(sheet1, no1, file1_col_fbc)) != int(get_cell_value(sheet2, no2, file2_col_fbc)):
                set_cell_color(sheet1, no1, file1_col_fbc, 'r')
                set_cell_comment(sheet1, no1, file1_col_fbc, txt + ' in DB is ' +
                                 str(int(get_cell_value(sheet2, no2, file2_col_fbc))), add=False)
                mark_row = True
            if int(get_cell_value(sheet1, no1, file1_col_card)) != int(get_cell_value(sheet2, no2, file2_col_card)):
                set_cell_color(sheet1, no1, file1_col_card, 'r')
                set_cell_comment(sheet1, no1, file1_col_card, txt + ' in DB is ' +
                                 str(int(get_cell_value(sheet2, no2, file2_col_card))), add=False)
                mark_row = True
            if int(get_cell_value(sheet1, no1, file1_col_channel)) != int(get_cell_value(sheet2, no2, file2_col_channel)):
                set_cell_color(sheet1, no1, file1_col_channel, 'r')
                set_cell_comment(sheet1, no1, file1_col_channel, txt + ' in DB is ' +
                                 str(int(get_cell_value(sheet2, no2, file2_col_channel))), add=False)
                mark_row = True
        if 'SL' in get_cell_value(sheet2, no2, file2_col_iotype):
            if get_cell_value(sheet1, no1, file1_col_modbus_address) is not None \
                    and get_cell_value(sheet1, no1, file2_col_modbus_address) is not None:
                if int(get_cell_value(sheet1, no1, file1_col_modbus_address)) != \
                        int(get_cell_value(sheet2, no2, file2_col_modbus_address)):
                    set_cell_color(sheet1, no1, file1_col_modbus_address, 'r')
                    set_cell_comment(sheet1, no1, file1_col_modbus_address, txt + ' in DB is ' +
                                     str(int(get_cell_value(sheet2, no2, file2_col_modbus_address))), add=False)
                    mark_row = True
            else:
                set_cell_color(sheet1, no1, file1_col_modbus_address, 'r')
                set_cell_comment(sheet1, no1, file1_col_modbus_address, txt + ' in DB is ' +
                                 str(int(get_cell_value(sheet2, no2, file2_col_modbus_address))), add=False)
                mark_row = True
            if 'D' in get_cell_value(sheet1, no2, file2_col_iotype):
                if get_cell_value(sheet1, no1, file1_col_bit) is not None \
                        and get_cell_value(sheet1, no1, file2_col_bit) is not None:
                    if int(get_cell_value(sheet1, no1, file1_col_bit)) != \
                            int(get_cell_value(sheet2, no2, file2_col_bit)):
                        set_cell_color(sheet1, no1, file1_col_bit, 'r')
                        set_cell_comment(sheet1, no1, file1_col_bit, txt + ' in DB is ' +
                                         str(int(get_cell_value(sheet2, no2, file2_col_bit))), add=False)
                        mark_row = True
                else:
                    set_cell_color(sheet1, no1, file1_col_bit, 'r')
                    set_cell_comment(sheet1, no1, file1_col_bit, txt + ' in DB is ' +
                                     str(int(get_cell_value(sheet2, no2, file2_col_bit))), add=False)
                    mark_row = True
            if 'A' in get_cell_value(sheet1, no2, file2_col_iotype):
                if get_cell_value(sheet1, no1, file1_col_gain) is not None \
                        and get_cell_value(sheet1, no1, file2_col_gain) is not None:
                    if float(get_cell_value(sheet1, no1, file1_col_gain)) != float(get_cell_value(sheet2, no2,
                                                                                                  file2_col_gain)):
                        set_cell_color(sheet1, no1, file1_col_gain, 'r')
                        set_cell_comment(sheet1, no1, file1_col_gain, txt + ' in DB is ' +
                                         str(get_cell_value(sheet2, no2, file2_col_gain)), add=False)
                        mark_row = True
                else:
                    set_cell_color(sheet1, no1, file1_col_gain, 'r')
                    set_cell_comment(sheet1, no1, file1_col_gain, txt + ' in DB is ' +
                                     str(get_cell_value(sheet2, no2, file2_col_gain)), add=False)
                    mark_row = True
            if get_cell_value(sheet1, no1, file1_col_slave) is not None \
                    and get_cell_value(sheet2, no2, file2_col_slave) is not None:
                if int(get_cell_value(sheet1, no1, file1_col_slave)) !=\
                        int(get_cell_value(sheet2, no2, file2_col_slave)):
                    set_cell_color(sheet1, no1, file1_col_slave, 'r')
                    set_cell_comment(sheet1, no1, file1_col_slave, txt + ' in DB is ' +
                                     str(get_cell_value(sheet2, no2, file2_col_slave)), add=False)
                    mark_row = True
            else:
                set_cell_color(sheet1, no1, file1_col_slave, 'r')
                set_cell_comment(sheet1, no1, file1_col_slave, txt + ' in DB is ' +
                                 str(get_cell_value(sheet2, no2, file2_col_slave)), add=False)
                mark_row = True
    if mark_row is True:
        set_cell_color(sheet1, no1, 1, 'r')


def compare_description(sheet1, no, sheet2, file1_col_loop, file2_col_tag, file1_col_description,
                        file2_col_description, max_rows2):
    """
    Function 'compare description' compares description between DB dump and CdA DB
    Parameters
    ----------
    sheet1 : workbook
        active sheet name DUMP from DNA
    no : int
        row number of searched row from CdA DB
    sheet2 : workbook
        active sheet name CdA DB

    file1_col_loop : int
        column number with TAG name in DUMP from DNA
    file2_col_tag : int
        column number with TAG name in DB from CdA
    file1_col_description : int
        column number with Description column in DB Dump from DNA
    file2_col_description : int
        column number with Description column in DB from CdA
    max_rows2
        numbers of rows in DB from CdA
    """
    found = False
    for i in range(2, max_rows2+1):
        if get_cell_value(sheet1, no, file1_col_loop) == get_cell_value(sheet2, i, file2_col_tag):
            found = True
            if get_cell_value(sheet1, no, file1_col_description) != get_cell_value(sheet2, i, file2_col_description):
                if str(get_cell_value(sheet1, no, file1_col_description)) not in \
                        str(get_cell_value(sheet2, i, file2_col_description)):
                    set_cell_color(sheet1, no, file1_col_description, 'r')
                    set_cell_comment(sheet1, no, file1_col_description, 'Other description' + ' in DB is in a row'
                                     + str(i) + ' (' + str(get_cell_value(sheet2, i, file2_col_tag)) + ')\n ' +
                                     str(get_cell_value(sheet2, i, file2_col_description)), add=False)
                    set_cell_color(sheet1, no, 1, 'r')
    if found is False:
        set_cell_color(sheet1, no, 1, 'g')
        set_cell_comment(sheet1, no, 1, 'Additional loop according to DB')

