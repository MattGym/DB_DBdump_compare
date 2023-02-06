# Program do porównywania db dumpa oraz bazy danych dostarczonej przez stocznię.
#tes

import openpyxl
# from openpyxl.comments import Comment
# from openpyxl.styles import PatternFill
# import pandas as pd

import tkinter
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox

from Functions import *


class CreateToolTip(object):
    def __init__(self, widget, text='widget info'):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.close)

    def enter(self, event=None):
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        # creates a toplevel window
        self.tw = tkinter.Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = tkinter.Label(self.tw, text=self.text, justify='left', relief='solid', borderwidth=1,
                              font=("times", "9", "normal"))
        label.pack(ipadx=1)

    def close(self, event=None):
        if self.tw:
            self.tw.destroy()


# Graphic animation function
root = Tk(className=' CdA DataBase Comparator')
root.geometry('725x185')
root.resizable(False, False)
# Variables
# File 1 Variables

file_path1 = ''
file_path_txt1 = StringVar()
file_path_txt1.set('Choose excel DB dump form DNA Explorer')
wb1 = openpyxl.Workbook()
active_sheet1 = wb1.active
max_rows1 = 0
max_col1 = 0

file1_col_tag = 0
file1_col_loop = 0
file1_col_package = 0
file1_col_description = 0
file1_col_min = 0
file1_col_max = 0
file1_col_unit = 0
file1_col_fbc = 0
file1_col_ibc = 0
file1_col_card = 0
file1_col_channel = 0
file1_col_instrument_code = 0
file1_col_signal_type = 0
file1_col_modbus_address = 0
file1_col_bit = 0
file1_col_gain = 0
file1_col_slave = 0
file1_col_link_signal_type = 0



# File 2 Variables
file_path2 = ''
file_path_txt2 = StringVar()
file_path_txt2.set('Choose appropriate DB form shipyard')
wb2 = openpyxl.Workbook()
active_sheet2 = wb2.active
max_rows2 = 0
max_col2 = 0

file2_col_tag = 0
file2_col_seq = 0
file2_col_iotype = 0
file2_col_package = 0
file2_col_description = 0
file2_col_min = 0
file2_col_max = 0
file2_col_unit = 0
file2_col_fbc = 0
file2_col_ibc = 0
file2_col_card = 0
file2_col_channel = 0
file2_col_instrument_code = 0
file2_col_modbus_address = 0
file2_col_bit = 0
file2_col_gain = 0
file2_col_slave = 0
file2_col_link_signal_type = 0


def analyze_file():
    global active_sheet1
    global max_rows1
    global max_col1
    global active_sheet2
    global max_rows2
    global max_col2
    wb1.active = wb1[sheet_choose1.get()]
    active_sheet1 = wb1.active
    max_rows1 = wb1.active.max_row
    max_col1 = wb1.active.max_column
    wb2.active = wb2[sheet_choose2.get()]
    active_sheet2 = wb2.active
    max_rows2 = wb2.active.max_row
    max_col2 = wb2.active.max_column
    fill_col_numbers()
    print(file_path1)
    print(file_path2)
    print('F1_col_modbusaddress', file1_col_modbus_address)
    print('F2_col_FBC', file2_col_fbc)
    if checkbox1_var.get() == 1:
        spaces1 = [file1_col_tag, file1_col_package, file1_col_min, file1_col_max, file1_col_fbc, file1_col_ibc,
                   file1_col_card, file1_col_channel, file1_col_instrument_code, file1_col_signal_type,
                   file1_col_modbus_address, file1_col_bit, file1_col_gain, file1_col_slave, file1_col_link_signal_type]
        numbers1 = [file1_col_min, file1_col_max, file1_col_fbc, file1_col_ibc, file1_col_card, file1_col_channel,
                    file1_col_modbus_address, file1_col_bit, file1_col_gain, file1_col_slave]
        string1 = []
        txt1 = [file1_col_tag, file1_col_loop, file1_col_package, file1_col_description, file1_col_unit,
                file1_col_signal_type, file1_col_package, file1_col_instrument_code, file1_col_link_signal_type]

        spaces2 = [file2_col_tag, file2_col_iotype, file2_col_package, file2_col_max, file2_col_fbc,
                   file2_col_ibc, file2_col_card, file2_col_channel, file2_col_instrument_code,
                   file2_col_modbus_address, file2_col_bit, file2_col_gain, file2_col_slave, file2_col_link_signal_type]
        numbers2 = [file2_col_seq, file2_col_max, file2_col_fbc, file2_col_ibc, file2_col_card, file2_col_channel,
                    file2_col_modbus_address, file2_col_bit, file2_col_gain, file2_col_slave]
        string2 = []
        txt2 = [file2_col_tag, file2_col_seq, file2_col_package, file2_col_description, file2_col_unit,
                file2_col_package,
                file2_col_instrument_code, file2_col_link_signal_type]
        remove_sign(active_sheet1, spaces1, max_rows1)
        variable_type_update(active_sheet1, numbers1, string1, txt1, max_rows1)
        wb1.save(file_path1)

        remove_sign(active_sheet2, spaces2, max_rows2)
        variable_type_update(active_sheet2, numbers2, string2, txt2, max_rows2)
        wb2.save(file_path2)

        if checkbox2_var.get() == 1:
            for i in range(2, max_rows2+1):
                no1 = search_row_in_dump(active_sheet1, active_sheet2, i, file1_col_tag, file2_col_tag, file2_col_seq,
                                         max_rows1)
                if no1 != 0:
                    compare_address(active_sheet1, active_sheet2, no1, i, file1_col_package, file2_col_package,
                                    file1_col_fbc, file2_col_fbc, file1_col_ibc, file2_col_ibc, file1_col_card,
                                    file2_col_card, file1_col_channel, file2_col_channel, file2_col_iotype,
                                    file1_col_modbus_address, file2_col_modbus_address, file1_col_bit, file2_col_bit,
                                    file1_col_gain, file2_col_gain, file1_col_slave, file2_col_slave)
                if no1 == 0 and 'TREAT' not in str(get_cell_value(active_sheet2, i, file2_col_iotype)):
                    set_cell_color(active_sheet2, i, 1, 'r')
                    set_cell_comment(active_sheet2, i, 1, 'Missing in DB')

        if checkbox3_var.get() == 1:
            for i in range(2, max_rows1+1):
                compare_description(active_sheet1, i, active_sheet2, file1_col_loop, file2_col_tag,
                                    file1_col_description, file2_col_description, max_rows2)

        print('saving 1st ')
        wb1.save(file_path1)
        print('saving 2nd')
        wb2.save(file_path2)
        print('Finish')


def fill_col_numbers():
    # DNA DUMP
    global file1_col_tag
    global  file1_col_loop
    global file1_col_package
    global file1_col_description
    global file1_col_min
    global file1_col_max
    global file1_col_unit
    global file1_col_fbc
    global file1_col_ibc
    global file1_col_card
    global file1_col_channel
    global file1_col_instrument_code
    global file1_col_signal_type
    global file1_col_modbus_address
    global file1_col_bit
    global file1_col_gain
    global file1_col_slave
    global file1_col_link_signal_type
    # CdA DB
    global file2_col_tag
    global file2_col_seq
    global file2_col_iotype
    global file2_col_package
    global file2_col_description
    global file2_col_min
    global file2_col_max
    global file2_col_unit
    global file2_col_fbc
    global file2_col_ibc
    global file2_col_card
    global file2_col_channel
    global file2_col_instrument_code
    global file2_col_modbus_address
    global file2_col_bit
    global file2_col_gain
    global file2_col_slave
    global file2_col_link_signal_type

    file1_col_tag = get_col_no(active_sheet1, '$(TAG)', max_col1)
    file1_col_loop = get_col_no(active_sheet1, '$(LOOP)', max_col1)
    file1_col_package = get_col_no(active_sheet1, '$(PACKAGE)', max_col1)
    file1_col_description = get_col_no(active_sheet1, '$(NAME)', max_col1)
    file1_col_min = get_col_no(active_sheet1, '$(MIN)', max_col1)
    file1_col_max = get_col_no(active_sheet1, '$(MAX)', max_col1)
    file1_col_unit = get_col_no(active_sheet1, '$(UNIT)', max_col1)
    file1_col_fbc = get_col_no(active_sheet1, '$(FBC)', max_col1)
    file1_col_ibc = get_col_no(active_sheet1, '$(IBC)', max_col1)
    file1_col_card = get_col_no(active_sheet1, '$(CARD)', max_col1)
    file1_col_channel = get_col_no(active_sheet1, '$(CHANNEL)', max_col1)
    file1_col_instrument_code = get_col_no(active_sheet1, '$(INSTRUMENT_CODE)', max_col1)
    file1_col_signal_type = get_col_no(active_sheet1, '$(TEMPLATE)', max_col1)
    file1_col_modbus_address = get_col_no(active_sheet1, '$(LIS_ADDR)', max_col1)
    file1_col_bit = get_col_no(active_sheet1, '$(LIS_BIT)', max_col1)
    file1_col_gain = get_col_no(active_sheet1, '$(LIS_GAIN)', max_col1)
    file1_col_slave = get_col_no(active_sheet1, '$(LIS_SLAVE)', max_col1)
    file1_col_link_signal_type = get_col_no(active_sheet1, '$(LIS_SIGNED)', max_col1)

    file2_col_tag = get_col_no(active_sheet2, 'loop', max_col2)
    file2_col_seq = get_col_no(active_sheet2, 'seq', max_col2)
    file2_col_iotype = get_col_no(active_sheet2, 'IO_type', max_col2)
    file2_col_package = get_col_no(active_sheet2, 'pcs', max_col2)
    file2_col_description = get_col_no(active_sheet2, 'design_loop', max_col2)
    file2_col_min = get_col_no(active_sheet2, 'min', max_col2)
    file2_col_max = get_col_no(active_sheet2, 'max', max_col2)
    file2_col_unit = get_col_no(active_sheet2, 'unit', max_col2)
    file2_col_fbc = get_col_no(active_sheet2, 'fbc', max_col2)
    file2_col_ibc = get_col_no(active_sheet2, 'rack_IBC', max_col2)
    file2_col_card = get_col_no(active_sheet2, 'card_slot', max_col2)
    file2_col_channel = get_col_no(active_sheet2, 'channel', max_col2)
    file2_col_instrument_code = get_col_no(active_sheet2, 'func', max_col2)
    file2_col_modbus_address = get_col_no(active_sheet2, 'modbus_address', max_col2)
    file2_col_bit = get_col_no(active_sheet2, 'bit_in_register', max_col2)
    file2_col_gain = get_col_no(active_sheet2, 'gain_factor', max_col2)
    file2_col_slave = get_col_no(active_sheet2, 'slave_no', max_col2)
    file2_col_link_signal_type = get_col_no(active_sheet2, 'signed_unsigned', max_col2)


def choose_file1():
    global file_path1
    global wb1
    root.filename = filedialog.askopenfilename(title='Choose file to open',
                                               filetypes=(('xlsx', '*.xlsx'), ('xls', '*.xls')))
    if len(root.filename) > 0:
        file_path1 = root.filename
        file_path_txt1.set('File: ' + file_path1)
    if len(file_path1) > 0:
        wb1 = openpyxl.load_workbook(file_path1)
        sheet_names1 = [wb1.sheetnames]
        sheet_choose1['values'] = tuple(sheet_names1[0])
        sheet_choose1.current(0)
        if len(file_path2) > 2:
            button_analyze['state'] = tkinter.NORMAL


def choose_file2():
    global file_path2
    global wb2
    root.filename = filedialog.askopenfilename(title='Choose file to open',
                                               filetypes=(('xlsx', '*.xlsx'), ('xls', '*.xls')))
    if len(root.filename) > 0:
        file_path2 = root.filename
        file_path_txt2.set('File: ' + file_path2)
    if len(file_path2) > 0:
        wb2 = openpyxl.load_workbook(file_path2)
        sheet_names2 = [wb2.sheetnames]
        sheet_choose2['values'] = tuple(sheet_names2[0])
        sheet_choose2.current(0)
        if len(file_path1) > 0:
            button_analyze['state'] = tkinter.NORMAL
# ------- Graphic user interface --------
# ---------------------------------------


file_label1 = Label(root, textvariable=file_path_txt1, width=60, anchor='w', relief='groove')
file_label1.place(x=10, y=20)
sheet_choose_select1 = tkinter.StringVar()
sheet_choose1 = ttk.Combobox(root, textvariable=sheet_choose_select1, width=10, height=1)
sheet_choose1.place(x=445, y=19)
button_select1 = Button(root, text='Select', command=choose_file1, height=1, width=10)
button_select1.place(x=540, y=17)

button_select2 = Button(root, text='Select', command=choose_file2, height=1, width=10)
button_select2.place(x=540, y=47)
file_label2 = Label(root, textvariable=file_path_txt2, width=60, anchor='w', relief='groove')

file_label2.place(x=10, y=49)
sheet_choose_select2 = tkinter.StringVar()
sheet_choose2 = ttk.Combobox(root, textvariable=sheet_choose_select2, width=10, height=1)
sheet_choose2.place(x=445, y=49)

button_analyze = Button(root, text='Analyze', command=analyze_file, height=3, width=10, state=tkinter.DISABLED)
button_analyze.place(x=635, y=17)

labelframe1 = ttk.Labelframe(root, width=705, height=95, labelanchor=NW, text='Check options')
labelframe1.place(x=10, y=80)

checkbox1_var = IntVar(root, 1)
checkbox1 = Checkbutton(root, text='Remove spaces and correct types', variable=checkbox1_var, onvalue=1,
                        offvalue=0, height=1, state=DISABLED)
checkbox1.place(x=20, y=95)
checkbox1_tt = CreateToolTip(checkbox1, "(Mandatory) removes spaces from cells and correcting datatypes.")

checkbox2_var = IntVar(root, 1)
checkbox2 = Checkbutton(root, text='Check IO HW and LINK address', variable=checkbox2_var, onvalue=1,
                        offvalue=0, height=1)
checkbox2.place(x=20, y=120)
checkbox2_tt = CreateToolTip(checkbox2, "Search for difference between DUMP and DB")

checkbox3_var = IntVar(root, 1)
checkbox3 = Checkbutton(root, text='Description check', variable=checkbox3_var, onvalue=1,
                        offvalue=0, height=1)
checkbox3.place(x=20, y=145)
checkbox3_tt = CreateToolTip(checkbox3, "Search for difference between DUMP and DB description")
root.mainloop()
